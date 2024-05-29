# -*- coding: utf-8 -*-
"""
Created on Tue Apr 30 10:14:08 2024

@author: Tan7080
"""



import pandas as pd
import numpy as np
import dateutil.parser as dparser
import glob
from datetime import datetime, timedelta
from openpyxl import load_workbook



def unhide_sheet(excel_file, sheet_name):
    try:
        wb = load_workbook(excel_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.sheet_state == 'hidden':
                ws.sheet_state = 'visible'  # Unhide the sheet
                wb.save(excel_file)
                print(f'Sheet "{sheet_name}" in file "{excel_file}" was hidden and has been unhidden.')
            else:
                print(f'Sheet "{sheet_name}" in file "{excel_file}" is already visible.')
        else:
            print(f'Sheet "{sheet_name}" not found in file "{excel_file}". Skipping...')
    except Exception as e:
        print(f'Error processing file "{excel_file}": {e}')

dh = []
name = 'PRODUCTION DPR Report'

for infile in glob.glob("*.xlsm"):
    # Unhide the sheet if it's hidden
    unhide_sheet(infile, name)
    df = pd.read_excel(infile, sheet_name=name)                    
                          
      
      
    def contains_area(cell):
      if isinstance(cell, str) and 'Area/Wells' in cell:
              return True
      else:
              return False
                
                # Use applymap() to check all cells in the DataFrame for the word 'ESP sensor data :'
    mask = df.applymap(contains_area)
    
            # Find the indices of all rows and columns that contain the word 'ESP sensor data :'
    rows, cols = np.where(mask)
    
    row= rows[0]
    col = cols[0]
    
    
    df1=df.iloc[row :,col:col+5]
    df1.columns = df1.iloc[0]
    df1 = df1[1:]
    df1=df1.iloc[:,[0,2,4]]
    df1= df1.rename(columns={'Area/Wells':'well_bore','Net oil differ, BO':'net_differ','Reason':'reason'})
    df1 =df1.dropna(subset='well_bore')
    sheet_date = dparser.parse(infile, fuzzy=True, dayfirst=True)

    # Subtract one day using timedelta
    previous_date = sheet_date - timedelta(days=1)
    
    # Format the previous date as YYYY-MM-DD
    previous_date_str = previous_date.strftime("%Y-%m-%d")
    
    df1.insert(0, "date", previous_date_str)
    
    dh.append(df1)
    print('{} done'.format(sheet_date))
    ''' concatinating the dataframes from all sheets and export as csv'''
        
        
       
dh=pd.concat(dh) 
   
dh.to_csv("dpr.csv")
        
print('finished safely')